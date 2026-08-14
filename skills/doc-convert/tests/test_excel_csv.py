"""excel_csv 的行为契约。

最重要的一条是 BOM：中文用户双击打开无 BOM 的 UTF-8 CSV，Excel 会按
本地代码页解码 → 满屏乱码。这不是锦上添花，是这条功能能不能用的分界线。
"""
import csv
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))
import excel_csv  # noqa: E402


def _make_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws.append(["姓名", "金额"])
    ws.append(["张三", 100])
    wb.save(str(path))


def test_xlsx_to_csv_writes_utf8_bom(tmp_path):
    src = tmp_path / "a.xlsx"
    _make_xlsx(src)
    dst = tmp_path / "a.csv"

    excel_csv.xlsx_to_csv(src, dst)

    raw = dst.read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf"), "缺 BOM 会让 Excel 打开中文 CSV 乱码"
    rows = list(csv.reader(dst.read_text(encoding="utf-8-sig").splitlines()))
    assert rows == [["姓名", "金额"], ["张三", "100"]]


def test_csv_to_xlsx_roundtrip(tmp_path):
    src = tmp_path / "b.csv"
    src.write_text("姓名,金额\n李四,200\n", encoding="utf-8-sig")
    dst = tmp_path / "b.xlsx"

    excel_csv.csv_to_xlsx(src, dst)

    ws = load_workbook(str(dst)).active
    # 2026-08-13 起 csv→xlsx 做保守数字推断："200" 这类数值格写成真数值，
    # 这是功能预期，不是回归。
    assert [[c.value for c in row] for row in ws.iter_rows()] == [
        ["姓名", "金额"],
        ["李四", 200],
    ]


def test_multi_sheet_requires_explicit_choice(tmp_path):
    src = tmp_path / "c.xlsx"
    wb = Workbook()
    wb.active.title = "一月"
    # 「二月」必须写点数据——评审后加固的空表护栏（见下方
    # test_empty_worksheet_refuses_to_export）会拒绝导出空工作表，这条测试要测
    # 的是"多表必须指定 --sheet"这另一条护栏，两者不能用同一张空表互相踩脚
    wb.create_sheet("二月").append(["二月数据"])
    wb.save(str(src))

    # 多表时静默只导第一张 = 用户丢数据还不知道。必须报错要求指定。
    with pytest.raises(SystemExit):
        excel_csv.xlsx_to_csv(src, tmp_path / "c.csv")

    excel_csv.xlsx_to_csv(src, tmp_path / "c.csv", sheet="二月")
    assert (tmp_path / "c.csv").is_file()


def test_empty_worksheet_refuses_to_export(tmp_path, capsys):
    """评审后加固：空工作表不应该静默产出 0 字节 CSV 还报"已生成"。

    这是四个脚本里唯一一处「静默产出空文件却报成功」，与本分支「拒绝时不产出
    文件」的纪律不一致——改成报错退出、不落盘。
    """
    src = tmp_path / "empty.xlsx"
    wb = Workbook()
    wb.save(str(src))  # 全新工作簿，工作表里没写过任何单元格
    dst = tmp_path / "empty.csv"

    with pytest.raises(SystemExit):
        excel_csv.xlsx_to_csv(src, dst)

    assert "一行数据都没有" in capsys.readouterr().err
    assert not dst.exists()


def test_xls_extension_gives_actionable_message(tmp_path, capsys):
    """.xls 是前端【Excel 文件】槽放行的旧格式，openpyxl 读不了；错误要指路。"""
    bad = tmp_path / "old.xls"
    bad.write_bytes(b"\xd0\xcf\x11\xe0")  # 假装是旧版二进制 xls 的文件头，内容不重要
    with pytest.raises(SystemExit):
        excel_csv.main([str(bad), "-o", str(tmp_path / "old.csv")])
    err = capsys.readouterr().err
    assert "旧格式" in err or "另存为" in err


def test_unknown_extension_exits_with_message(tmp_path, capsys):
    bad = tmp_path / "d.txt"
    bad.write_text("x", encoding="utf-8")
    with pytest.raises(SystemExit):
        excel_csv.main([str(bad), "-o", str(tmp_path / "d.csv")])
    assert "只支持" in capsys.readouterr().err


def _make_formula_xlsx(path: Path) -> None:
    """程序生成的 xlsx：公式写进去了，但没有任何缓存计算结果。

    这不是刻意构造的极端样本——openpyxl（含本技能自己的 rows_to_xlsx.py）、
    以及各类系统导出的 xlsx 全是这样，因为写文件的程序里没有公式引擎。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws.append(["项目", "金额"])
    ws.append(["甲", 100])
    ws.append(["乙", 200])
    ws.append(["合计", "=SUM(B2:B3)"])
    wb.save(str(path))


def test_uncached_formula_refuses_export(tmp_path, capsys):
    """复审实测：data_only=True 读的是"上次保存时缓存的结果"，程序生成的 xlsx
    根本没有缓存值 → 整列合计读出来是 None → 导出的 CSV 里那格是空的，脚本却
    exit 0 报「已生成」。用户拿着一份缺了汇总列的表去汇报，发现时已经晚了。
    """
    src = tmp_path / "f.xlsx"
    _make_formula_xlsx(src)
    dst = tmp_path / "f.csv"

    with pytest.raises(SystemExit) as e:
        excel_csv.xlsx_to_csv(src, dst)

    assert e.value.code != 0
    err = capsys.readouterr().err
    assert "公式" in err
    assert "--allow-empty-formulas" in err  # 报错必须给出"怎么强行继续"
    assert not dst.exists()  # 拒绝时不产出文件


def test_uncached_formula_exports_when_explicitly_allowed(tmp_path):
    """门禁的另一半：用户明确知情后可以照转，那几格空着。"""
    src = tmp_path / "f2.xlsx"
    _make_formula_xlsx(src)
    dst = tmp_path / "f2.csv"

    excel_csv.xlsx_to_csv(src, dst, allow_empty_formulas=True)

    rows = list(csv.reader(dst.read_text(encoding="utf-8-sig").splitlines()))
    assert rows[-1] == ["合计", ""]


def test_formula_gate_does_not_block_normal_files(tmp_path):
    """门禁不能误伤：没有公式的普通表格照转不误。"""
    src = tmp_path / "plain.xlsx"
    _make_xlsx(src)
    dst = tmp_path / "plain.csv"

    excel_csv.xlsx_to_csv(src, dst)

    assert dst.is_file()


def test_empty_csv_refuses_to_produce_xlsx(tmp_path, capsys):
    """复审实测：反方向（xlsx→csv）已经拦了空表，这一头漏了——空 CSV 会产出
    一个只有空工作表的 xlsx 还报「已生成」，用户打开一片空白只会以为转丢了。
    """
    src = tmp_path / "e.csv"
    src.write_text("", encoding="utf-8")
    dst = tmp_path / "e.xlsx"

    with pytest.raises(SystemExit) as e:
        excel_csv.csv_to_xlsx(src, dst)

    assert e.value.code != 0
    assert "一行数据都没有" in capsys.readouterr().err
    assert not dst.exists()


def test_main_wraps_unexpected_error_in_chinese(tmp_path, monkeypatch, capsys):
    """main() 的兜底层：写盘失败 / 损坏的 zip 包不能漏英文堆栈。"""
    src = tmp_path / "a.xlsx"
    _make_xlsx(src)

    def _boom(*a, **k):
        raise OSError(30, "Read-only file system")

    monkeypatch.setattr(excel_csv, "xlsx_to_csv", _boom)

    with pytest.raises(SystemExit) as e:
        excel_csv.main([str(src), "-o", str(tmp_path / "a.csv")])

    assert e.value.code != 0
    err = capsys.readouterr().err
    assert "处理过程中出错" in err
    assert "Traceback" not in err


def test_csv_to_xlsx_infers_numbers_conservatively(tmp_path, capsys):
    """csv→xlsx 数字推断：数值写成真数值（=SUM 能算），编号样的坚决保文本。
    护栏三条：前导零、纯整数位数 ≥10、含非数值字符。"""
    src = tmp_path / "n.csv"
    src.write_text(
        "项目,金额,发票号,电话,备注\n"
        "差旅,\"1,200.50\",24312000000123456789,13800138000,正常\n"
        "餐饮,-56,007,1.5E+3,0.5\n",
        encoding="utf-8-sig",
    )
    dst = tmp_path / "n.xlsx"
    assert excel_csv.main([str(src), "-o", str(dst)]) == 0
    ws = load_workbook(dst).active
    # 数值列：千分位被剥掉、负数、小数都成真数值
    assert ws["B2"].value == 1200.5 and ws["B2"].data_type == "n"
    assert ws["B3"].value == -56 and ws["B3"].data_type == "n"
    assert ws["E3"].value == 0.5
    # 编号护栏：20 位发票号、11 位手机号、前导零、科学计数法——全部保文本
    assert ws["C2"].value == "24312000000123456789"
    assert ws["D2"].value == "13800138000"
    assert ws["C3"].value == "007"
    assert ws["D3"].value == "1.5E+3"
    # 表头行是普通文本
    assert ws["B1"].value == "金额"
    # 透明度：转换报告点名列
    out = capsys.readouterr().out
    assert "写成真数值" in out
    assert "保留为文本" in out


def test_coerce_cell_edge_cases():
    assert excel_csv._coerce_cell("128.5") == (128.5, "num")
    assert excel_csv._coerce_cell("1,200.50") == (1200.5, "num")
    assert excel_csv._coerce_cell("-42") == (-42, "num")
    assert excel_csv._coerce_cell("999999999") == (999999999, "num")     # 9 位，转
    assert excel_csv._coerce_cell("1234567890")[1] == "guarded"          # 10 位，保
    assert excel_csv._coerce_cell("0") == (0, "num")
    assert excel_csv._coerce_cell("0.5") == (0.5, "num")
    assert excel_csv._coerce_cell("007")[1] == "guarded"
    assert excel_csv._coerce_cell("12345678901.5") == (12345678901.5, "num")  # 12 位有效数字，够精度
    # 终审留任项收口（2026-08-14）：带小数的数字总位数超 15 位同样会被
    # float/Excel 的精度静默截断（12345678901234567890.5 会变 1.234...e+19），
    # 与纯整数护栏是同一个坑的另一条进口，一并 guarded
    assert excel_csv._coerce_cell("12345678901234567890.5")[1] == "guarded"
    assert excel_csv._coerce_cell("123456789012345.5")[1] == "guarded"        # 16 位，超
    assert excel_csv._coerce_cell("1234567890123.45") == (1234567890123.45, "num")  # 恰 15 位，转
    assert excel_csv._coerce_cell("1,23")[1] == "text"   # 假千分位
    assert excel_csv._coerce_cell("abc")[1] == "text"
    assert excel_csv._coerce_cell("")[1] == "text"
    assert excel_csv._coerce_cell(" 42 ") == (42, "num")  # 首尾空白剥掉再判
    # M-2：全角数字不是本技能要推断的"数值"——\d 默认 Unicode 感知会连
    # 全角一起吃掉，转成 ASCII 数值会悄悄改写字形，必须保文本
    assert excel_csv._coerce_cell("１２３")[1] == "text"
