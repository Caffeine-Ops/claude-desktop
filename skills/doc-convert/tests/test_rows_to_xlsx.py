"""rows_to_xlsx 的行为契约 —— 本 PR 全部质量门禁的落点。

四条最要紧的：
  1. 数字要写成真数值，不然用户 =SUM() 得 0（PR 1 的 csv→xlsx 已经踩过一次，
     那次只能靠文档提醒补救；这次是我们自己生成，没有借口）。
  2. 「看不清」和「本来就没有」必须区分：都标成无法识别会制造大量假警报，
     用户三天就学会无视所有黄格子，标记随之失效。
  3. 模型偷偷加列要当场拒绝——表头是跟用户确认过的契约。
  4. 识别率不到一半直接拒绝出文件：一份一半是问号的台账，用户核对的工夫
     比自己录还多。
"""
import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

SCRIPTS = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS))
import rows_to_xlsx  # noqa: E402


def _write_json(path: Path, payload: dict) -> Path:
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return path


def _run(src: Path, out: Path, *extra: str):
    return subprocess.run(
        [sys.executable, str(SCRIPTS / "rows_to_xlsx.py"), str(src), "-o", str(out), *extra],
        capture_output=True, text=True,
    )


def test_numbers_are_written_as_numbers(tmp_path):
    src = _write_json(tmp_path / "a.json", {
        "headers": ["项目", "金额"],
        "rows": [{"项目": "差旅", "金额": 128.5}],
    })
    out = tmp_path / "a.xlsx"
    assert _run(src, out).returncode == 0
    ws = load_workbook(out).active
    assert ws["B2"].value == 128.5
    assert ws["B2"].data_type == "n", "写成文本的话用户 =SUM() 会得 0"


def test_uncertain_cell_is_marked_and_filled_yellow(tmp_path):
    src = _write_json(tmp_path / "b.json", {
        "headers": ["项目", "金额"],
        "rows": [{"项目": "餐饮", "金额": None,
                  "_存疑": [{"字段": "金额", "原因": "折痕遮挡"}],
                  "_来源": "IMG_1.jpg"}],
    })
    out = tmp_path / "b.xlsx"
    assert _run(src, out).returncode == 0
    wb = load_workbook(out)
    ws = wb.active
    assert ws["B2"].value == rows_to_xlsx.UNREADABLE_TEXT
    assert rows_to_xlsx.YELLOW_HEX in str(ws["B2"].fill.start_color.rgb)
    review = wb["待核对"]
    assert [c.value for c in review[2]] == ["IMG_1.jpg", "金额", "折痕遮挡"]


def test_missing_but_not_uncertain_stays_empty(tmp_path):
    """票据上本来就没有税额 ≠ 看不清税额。混为一谈会制造假警报。"""
    src = _write_json(tmp_path / "c.json", {
        "headers": ["项目", "税额"],
        "rows": [{"项目": "打车", "税额": None}],
    })
    out = tmp_path / "c.xlsx"
    assert _run(src, out).returncode == 0
    wb = load_workbook(out)
    assert wb.active["B2"].value is None
    assert "待核对" not in wb.sheetnames, "一个存疑都没有时不该多出一张空表"


def test_unknown_field_is_refused_with_row_number(tmp_path):
    src = _write_json(tmp_path / "d.json", {
        "headers": ["项目"],
        "rows": [{"项目": "住宿"}, {"项目": "机票", "偷加的列": "x"}],
    })
    out = tmp_path / "d.xlsx"
    proc = _run(src, out)
    assert proc.returncode != 0
    assert "第 2 行" in proc.stderr and "偷加的列" in proc.stderr
    assert not out.exists()


def test_over_half_uncertain_is_refused(tmp_path):
    src = _write_json(tmp_path / "e.json", {
        "headers": ["日期", "金额"],
        "rows": [{"日期": None, "金额": None,
                  "_存疑": [{"字段": "日期", "原因": "模糊"},
                            {"字段": "金额", "原因": "模糊"}]}],
    })
    out = tmp_path / "e.xlsx"
    proc = _run(src, out)
    assert proc.returncode != 0
    assert "超过一半" in proc.stderr
    assert not out.exists()


def test_empty_rows_is_refused(tmp_path):
    src = _write_json(tmp_path / "f.json", {"headers": ["项目"], "rows": []})
    out = tmp_path / "f.xlsx"
    proc = _run(src, out)
    assert proc.returncode != 0
    assert not out.exists()


def test_jsonl_without_headers_is_refused(tmp_path):
    src = tmp_path / "g.jsonl"
    src.write_text('{"项目": "住宿"}\n', encoding="utf-8")
    out = tmp_path / "g.xlsx"
    proc = _run(src, out)
    assert proc.returncode != 0
    assert "--headers" in proc.stderr
    assert not out.exists()


def test_jsonl_with_headers_works(tmp_path):
    src = tmp_path / "h.jsonl"
    src.write_text('{"项目": "住宿", "金额": 300}\n{"项目": "机票", "金额": 1200}\n',
                   encoding="utf-8")
    out = tmp_path / "h.xlsx"
    assert _run(src, out, "--headers", "项目", "金额").returncode == 0
    ws = load_workbook(out).active
    assert ws["A3"].value == "机票"


def test_boolean_is_written_as_text_not_number(tmp_path):
    """Python 里 bool 是 int 的子类，不特判会把 True 写成 1。"""
    src = _write_json(tmp_path / "i.json", {
        "headers": ["项目", "已报销"],
        "rows": [{"项目": "打车", "已报销": True}],
    })
    out = tmp_path / "i.xlsx"
    assert _run(src, out).returncode == 0
    assert load_workbook(out).active["B2"].value == "True"


# --- 补充：main() 的兜底 except Exception —— Task 2 评审升级的全局约束。
# 只测 _die 路径测不到兜底本身，因为 _die 抛的是 SystemExit（继承
# BaseException 不是 Exception），到不了外层的 except Exception；必须让
# 内部函数抛一个普通异常才能验证兜底真的接住了。同 pdf_tables.py /
# pdf_render.py / doc_text.py 已落地的写法。

def test_main_wraps_unexpected_exception(tmp_path, monkeypatch, capsys):
    src = _write_json(tmp_path / "j.json", {
        "headers": ["项目"],
        "rows": [{"项目": "住宿"}],
    })
    out = tmp_path / "j.xlsx"

    def _boom(*a, **kw):
        raise RuntimeError("boom")

    monkeypatch.setattr(rows_to_xlsx, "build", _boom)
    with pytest.raises(SystemExit) as exc:
        rows_to_xlsx.main([str(src), "-o", str(out)])
    assert exc.value.code != 0
    err = capsys.readouterr().err
    assert "[doc-convert] 错误：" in err
    # 注意：这里不断言 "Traceback" not in err——这个测试是进程内直接调
    # main()，pytest 的 capsys 只会捕获 print() 写进 stderr 的内容，未捕获
    # 的异常会被 pytest 自己的 traceback 渲染逻辑接管、根本不会进 capsys。
    # 也就是说即便没有 except Exception 兜底，这条断言在这里也会恒真——是
    # 装饰性的、验证不了任何东西（评审指出的 Minor）。真正跨进程、能验证
    # 子进程 stderr 里确实没有裸 Traceback 的测试见下面
    # test_main_wraps_unexpected_exception_in_real_subprocess。


# --- 评审 Important 补充：以下 6 条是 Task 6 评审（2026-08-11）实测发现的
# 边界，任务书本身没覆盖，但都落在本文件「宁可不产出，不产出看似正常实则
# 有缺陷的文件」的职责范围内。

def test_uncertain_field_not_in_headers_is_refused(tmp_path):
    """[I-1] _存疑 里的字段名如果和表头对不上，build() 里 `name in unc` 的
    交集判断会静默找不到匹配——那一格不染黄、不写「⚠ 无法识别」、也不计入
    50% 门禁，模型明明说了"这格没看清"，用户拿到的却是一个完全正常的格子。
    必须在 validate() 阶段就当场拦住，同「偷加列」一个形状。
    """
    src = _write_json(tmp_path / "k.json", {
        "headers": ["项目", "金额"],
        "rows": [{"项目": "餐饮", "金额": 1,
                  "_存疑": [{"字段": "根本没这列", "原因": "糊"}],
                  "_来源": "x.jpg"}],
    })
    out = tmp_path / "k.xlsx"
    proc = _run(src, out)
    assert proc.returncode != 0
    assert proc.stderr.startswith("[doc-convert] 错误：")
    assert "根本没这列" in proc.stderr
    assert "表头完全一致" in proc.stderr, "错误信息要指导用户怎么自救"
    assert not out.exists()


def test_top_level_json_array_is_refused(tmp_path):
    """[I-2a] 模型把 rows 数组直接当成整个文件是很自然的手滑；不提前拦，
    payload.get("headers") 会在 list 上抛 AttributeError，被兜底接住后
    吐出一句用户看不懂的英文异常。"""
    src = tmp_path / "l.json"
    src.write_text(json.dumps([{"项目": "住宿"}], ensure_ascii=False), encoding="utf-8")
    out = tmp_path / "l.xlsx"
    proc = _run(src, out)
    assert proc.returncode != 0
    assert proc.stderr.startswith("[doc-convert] 错误：")
    assert "AttributeError" not in proc.stderr
    assert "顶层" in proc.stderr, "错误信息要指导用户怎么自救"
    assert not out.exists()


def test_jsonl_line_that_is_not_an_object_is_refused(tmp_path):
    """[I-2b] .jsonl 某一行是合法 JSON 但不是对象（比如手滑写成了数组），
    要带着行号当场拦住，而不是让它滚到后面变成一句看不懂的英文异常。"""
    src = tmp_path / "m.jsonl"
    src.write_text("[1,2,3]\n", encoding="utf-8")
    out = tmp_path / "m.xlsx"
    proc = _run(src, out, "--headers", "项目")
    assert proc.returncode != 0
    assert proc.stderr.startswith("[doc-convert] 错误：")
    assert "第 1 行" in proc.stderr
    assert "对象" in proc.stderr, "错误信息要指导用户怎么自救"
    assert not out.exists()


def test_row_that_is_a_bare_string_is_refused(tmp_path):
    """[I-2c] rows 里混进一个非对象元素（比如一个裸字符串）会让
    `for k in row` 逐字符遍历那个字符串，报出的"多余字段"是一堆断成单字的
    假信息，完全误导用户去哪查。必须在真正遍历字段之前先判类型。"""
    src = _write_json(tmp_path / "n.json", {
        "headers": ["项目"],
        "rows": ["我不是对象"],
    })
    out = tmp_path / "n.xlsx"
    proc = _run(src, out)
    assert proc.returncode != 0
    assert proc.stderr.startswith("[doc-convert] 错误：")
    assert "第 1 行" in proc.stderr
    assert "对象" in proc.stderr, "错误信息要指导用户怎么自救"
    assert not out.exists()


def test_sheet_name_with_illegal_character_is_refused(tmp_path):
    """[I-3] Excel 表名禁用 / \\ [ ] : * ?，模型给一个带日期斜杠的表名
    （如 "2026/03 台账"）太正常了。不提前拦，openpyxl 会抛英文 ValueError，
    而且 build() 在 mkdir/save 之前跑，落盘前就该拒绝，不留半成品文件。"""
    src = _write_json(tmp_path / "o.json", {
        "headers": ["项目"],
        "rows": [{"项目": "住宿"}],
    })
    out = tmp_path / "o.xlsx"
    proc = _run(src, out, "--sheet", "2026/03 台账")
    assert proc.returncode != 0
    assert proc.stderr.startswith("[doc-convert] 错误：")
    assert "不允许的字符" in proc.stderr, "错误信息要指导用户怎么自救"
    assert not out.exists()


def test_string_starting_with_equals_is_written_as_text_not_formula(tmp_path):
    """[I-4] openpyxl 见到以 = 开头的字符串会自动当公式写入，直接违反本
    文件 docstring 自己写的契约（"输出字符串就写成文本"）。评审已用存盘
    往返验证过这个修法：data_type 稳定为 's'，值原样保留。"""
    src = _write_json(tmp_path / "p.json", {
        "headers": ["备注"],
        "rows": [{"备注": "=1+1"}],
    })
    out = tmp_path / "p.xlsx"
    assert _run(src, out).returncode == 0
    ws = load_workbook(out).active
    assert ws["A2"].value == "=1+1"
    assert ws["A2"].data_type == "s"


def test_sheet_name_colliding_with_review_sheet_is_refused(tmp_path):
    """[Minor 5] `--sheet 待核对` 会撞上存疑核对表的保留名——评审实测 openpyxl
    自动改名成 `['待核对', '待核对1']`，结果数据表反而叫「待核对」，评审表叫
    「待核对1」，两张表意思对调，比没有提示更糊涂。必须在写盘前当场拒绝。"""
    src = _write_json(tmp_path / "r.json", {
        "headers": ["项目", "金额"],
        "rows": [{"项目": "餐饮", "金额": 1,
                  "_存疑": [{"字段": "金额", "原因": "糊"}], "_来源": "x.jpg"}],
    })
    out = tmp_path / "r.xlsx"
    proc = _run(src, out, "--sheet", rows_to_xlsx.REVIEW_SHEET)
    assert proc.returncode != 0
    assert proc.stderr.startswith("[doc-convert] 错误：")
    assert "保留" in proc.stderr, "错误信息要指导用户怎么自救"
    assert not out.exists()


def test_main_wraps_unexpected_exception_in_real_subprocess(tmp_path):
    """[Minor] test_main_wraps_unexpected_exception 的 "Traceback not in
    err" 断言是装饰性的（见上面那条测试新加的注释）：那条测试进程内直接调
    main()，未捕获异常根本进不了 capsys 捕获的 stderr。这里换一条真正跨
    进程的测试，用一个 I-1~I-4 没有专门覆盖、但确实会在 build() 里炸出
    原生 AttributeError 的输入（meta 不是对象，`meta.get("标题")` 在字符串
    上调用），验证子进程的 stderr 里真的看不到裸 Traceback——这才是「不让
    Python 堆栈漏到用户面前」这条全局约束真正要保证的场景。
    """
    src = _write_json(tmp_path / "q.json", {
        "headers": ["项目"],
        "meta": "不是对象",
        "rows": [{"项目": "住宿"}],
    })
    out = tmp_path / "q.xlsx"
    proc = _run(src, out)
    assert proc.returncode != 0
    assert "[doc-convert] 错误：" in proc.stderr
    assert "Traceback" not in proc.stderr
    assert not out.exists()
