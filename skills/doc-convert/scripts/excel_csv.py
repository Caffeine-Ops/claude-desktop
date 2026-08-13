#!/usr/bin/env python3
"""Excel ↔ CSV 双向转换。

**刻意不用 pandas**：这件事用内置 csv + openpyxl 就够了，而 pandas 连同
numpy 约 84 MB，比本技能其余依赖加起来还大。为一次读写背这个包不划算。

两个非显然的决定，都是「不这么做用户就会踩坑」：

1. **写 CSV 一律带 UTF-8 BOM**（encoding="utf-8-sig"）。中文用户双击打开
   无 BOM 的 UTF-8 CSV 时，Excel 按本地代码页解码 → 满屏乱码，然后用户
   会认为是我们转错了。BOM 让 Excel 认出编码。读 CSV 同样用 utf-8-sig，
   它对没有 BOM 的文件也能正常工作（只在有 BOM 时吃掉它）。
2. **多工作表时拒绝猜**。静默只导第一张表 = 用户丢了数据还不知道。宁可
   报错要求他指定 --sheet。
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

_XLSX_SUFFIXES = {".xlsx", ".xlsm"}
_CSV_SUFFIXES = {".csv"}


def _die(msg: str) -> None:
    """统一的中文报错出口，措辞与 doc_text.py / pdf_tables.py 等脚本对齐。"""
    print(f"[doc-convert] 错误：{msg}", file=sys.stderr)
    raise SystemExit(2)


# csv→xlsx 数字推断的三条护栏（顺序即优先级）。设计取舍见 SKILL.md B3：
# 编号（发票号/手机号/身份证号）长得像数字但不是数值——转成 number 会被
# Excel 的 15 位精度静默截断成科学计数法，正是本技能要拦的「看起来正常
# 实则数字有缺陷」。宁可把十亿级无小数点的大金额保守留成文本（用户在
# Excel 里一步能改回数值），也不冒截断编号的险。
# re.ASCII：不加的话 \d 是 Unicode 感知的，会把全角数字（１２００）也当数字
# 匹配上，int()/float() 再把它转成 ASCII 数值——字形被悄悄改写。全角数字
# 不该被当成本技能要推断的"数值"，一律保留成文本原样。
_THOUSANDS_RE = re.compile(r"^[+-]?\d{1,3}(,\d{3})+(\.\d+)?$", re.ASCII)
_PLAIN_NUM_RE = re.compile(r"^[+-]?\d+(\.\d+)?$", re.ASCII)
_MAX_INT_DIGITS = 9  # 纯整数位数 ≥10（手机号 11 位起）一律保文本


def _coerce_cell(text: str) -> tuple[int | float | str, str]:
    """CSV 单元格 → (写入值, 标签)。标签："num" 转成数值 / "guarded" 长得像
    数字但被护栏留成文本 / "text" 普通文本。"""
    s = text.strip()
    if not s:
        return text, "text"
    if _THOUSANDS_RE.match(s):
        s = s.replace(",", "")  # 千分位只是显示形式，剥掉不算改值（同 A2 规则）
    elif not _PLAIN_NUM_RE.match(s):
        return text, "text"
    int_part = s.lstrip("+-").split(".", 1)[0]
    if len(int_part) > 1 and int_part.startswith("0"):
        return text, "guarded"  # 前导零：区号/编号特征（"0"、"0.5" 不算）
    if "." not in s:
        if len(int_part) > _MAX_INT_DIGITS:
            return text, "guarded"
        return int(s), "num"
    return float(s), "num"


def _uncached_formula_cells(src: Path, sheet_name: str) -> list[str]:
    """找出「是公式、但没有缓存计算结果」的单元格坐标。

    评审实测：`load_workbook(data_only=True)` 的真实语义是「读 Excel 上次保存
    时**缓存**下来的计算结果」，而不是「现在算一遍」——openpyxl 不含公式引擎。
    Excel / WPS 自己保存的文件都带缓存值，所以日常文件没事；但**程序生成的
    xlsx（包括 openpyxl 自己写出来的）没有缓存值**，这些格子读出来全是 None，
    到下游变成空字符串。后果是一整列「合计 / 小计 / 同比」凭空消失，脚本还
    exit 0 打印「已生成」——用户拿着一份缺了汇总列的表去汇报，发现时已经晚了。
    这正是本分支「宁可不产出，也不产出一份看起来正常实则有缺陷的文件」要拦的。

    判据用 `cell.data_type == "f"` 而不是 `str(value).startswith("=")`：前者是
    openpyxl 对单元格类型的正式标记，普通数组公式（ArrayFormula 对象，不是
    字符串）也能覆盖到。
    """
    wb = load_workbook(str(src), data_only=False)
    if sheet_name not in wb.sheetnames:
        return []
    ws_formula = wb[sheet_name]
    wb_value = load_workbook(str(src), data_only=True)
    ws_value = wb_value[sheet_name]

    bad: list[str] = []
    for row_f, row_v in zip(ws_formula.iter_rows(), ws_value.iter_rows()):
        for cell_f, cell_v in zip(row_f, row_v):
            if cell_f.data_type == "f" and cell_v.value is None:
                bad.append(cell_f.coordinate)
    return bad


def xlsx_to_csv(
    src: Path, dst: Path, sheet: str | None = None, allow_empty_formulas: bool = False
) -> None:
    wb = load_workbook(str(src), data_only=True)
    names = wb.sheetnames
    if sheet is None:
        if len(names) > 1:
            print(
                f"[doc-convert] 错误：{src.name} 有 {len(names)} 张工作表"
                f"（{'、'.join(names)}），请用 --sheet 指定要导出哪一张。",
                file=sys.stderr,
            )
            raise SystemExit(2)
        ws = wb[names[0]]
    else:
        if sheet not in names:
            print(
                f"[doc-convert] 错误：没有名为「{sheet}」的工作表。"
                f"可选：{'、'.join(names)}",
                file=sys.stderr,
            )
            raise SystemExit(2)
        ws = wb[sheet]

    # 评审后加固：先把行读进内存判断有没有数据，再决定要不要落盘。原实现是边读
    # 边写，源表一行数据都没有时会在磁盘上留下一个 0 字节的 CSV、还打印"已生成"——
    # 这是四个脚本里唯一一处"静默产出空文件却报成功"，跟本分支「宁可不产出，也
    # 不产出一份看起来正常实则有缺陷的文件」的纪律不符。0 字节文件用户双击打开
    # 会一脸问号，还以为源表本来就有数据、是转换弄丢的。
    rows = [
        ["" if v is None else v for v in row] for row in ws.iter_rows(values_only=True)
    ]
    if not rows:
        print(
            f"[doc-convert] 错误：{src.name}"
            + (f" 的工作表「{ws.title}」" if sheet else "")
            + " 里一行数据都没有，没有内容可导出。",
            file=sys.stderr,
        )
        raise SystemExit(2)

    # 公式列门禁。做成「默认拒绝 + 显式开关」而不是「转完加一行提示」，理由同
    # docx_to_pdf.py 的 --allow-textonly：提示会被模型和用户一起略过，而这份
    # CSV 看起来完全正常（就是少了几个格子），带着它去汇报没人会当场发现。
    # 降级本身可以接受，不知情的降级不行。
    if not allow_empty_formulas:
        bad = _uncached_formula_cells(src, ws.title)
        if bad:
            shown = "、".join(bad[:8]) + ("…" if len(bad) > 8 else "")
            print(
                f"[doc-convert] 错误：{src.name} 的工作表「{ws.title}」里有 "
                f"{len(bad)} 个公式单元格没有保存计算结果（{shown}），"
                "直接导出的话这些格子在 CSV 里会是空的（常见于程序生成的 xlsx，"
                "Excel 自己保存的文件不会这样）。\n"
                "  · 想保住这些数：用 Excel / WPS 打开这份文件、再另存一次让公式"
                "算出结果，然后重试。\n"
                "  · 确认这些空格无所谓：重跑时加 --allow-empty-formulas。",
                file=sys.stderr,
            )
            raise SystemExit(2)

    dst.parent.mkdir(parents=True, exist_ok=True)
    # newline="" 是 csv 模块的硬要求，不写会在 Windows 上多出空行
    with dst.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def csv_to_xlsx(src: Path, dst: Path) -> None:
    # 复审实测：反方向（xlsx→csv）上面已经拦了"源表一行都没有"，这一头漏了。
    # 空 CSV 进来会产出一个 4.7 KB、只有一张空工作表的 xlsx 还打印「已生成」，
    # 用户双击打开一片空白，只会以为是转换弄丢了数据。两个方向的纪律要一致：
    # 先读进内存判断有没有内容，再决定要不要落盘。
    with src.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        print(
            f"[doc-convert] 错误：{src.name} 里一行数据都没有，没有内容可转换。",
            file=sys.stderr,
        )
        raise SystemExit(2)

    wb = Workbook()
    ws = wb.active
    num_cols: dict[int, int] = {}      # 列号 → 转成数值的格数
    guarded_cols: dict[int, int] = {}  # 列号 → 被护栏保成文本的格数
    for row in rows:
        out_row = []
        for col, cell in enumerate(row, start=1):
            value, tag = _coerce_cell(cell)
            if tag == "num":
                num_cols[col] = num_cols.get(col, 0) + 1
            elif tag == "guarded":
                guarded_cols[col] = guarded_cols.get(col, 0) + 1
            out_row.append(value)
        ws.append(out_row)
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dst))

    if num_cols:
        cols = "、".join(get_column_letter(c) for c in sorted(num_cols))
        print(f"[doc-convert] 已把 {sum(num_cols.values())} 个数字单元格"
              f"写成真数值（第 {cols} 列），可以直接用 =SUM() 等公式计算。")
    if guarded_cols:
        cols = "、".join(get_column_letter(c) for c in sorted(guarded_cols))
        print(f"[doc-convert] 另有 {sum(guarded_cols.values())} 个疑似编号的"
              f"单元格（第 {cols} 列，前导零或位数过长）保留为文本，避免被 "
              "Excel 按 15 位精度截断。确实需要按数值计算的话，"
              "在 Excel 里选中该列改成数值格式即可。")


def main(argv: list[str] | None = None) -> int:
    try:
        ap = argparse.ArgumentParser(
            description="Excel 与 CSV 互转（方向按输入扩展名自动判定）"
        )
        ap.add_argument("input", help="输入 .xlsx / .xlsm / .csv 文件")
        ap.add_argument("-o", "--output", required=True, help="输出文件")
        ap.add_argument("--sheet", default=None, help="仅 xlsx→csv：指定工作表名")
        ap.add_argument(
            "--allow-empty-formulas",
            action="store_true",
            help="仅 xlsx→csv：允许导出没有缓存计算结果的公式单元格（这些格子会是空的）",
        )
        args = ap.parse_args(argv)

        src = Path(args.input)
        if not src.is_file():
            _die(f"找不到输入文件 {src}")

        suffix = src.suffix.lower()
        dst = Path(args.output)
        if suffix in _XLSX_SUFFIXES:
            xlsx_to_csv(src, dst, args.sheet, args.allow_empty_formulas)
        elif suffix in _CSV_SUFFIXES:
            csv_to_xlsx(src, dst)
        elif suffix == ".xls":
            # 评审后加固：前端【Excel 文件】槽放行 .xls，但 openpyxl 只认 .xlsx/.xlsm
            # 这种 zip 容器格式，读不了 .xls 的老二进制格式。单独分支给出"下一步该
            # 干嘛"，而不是让它落进下面那条"只支持..."的通用提示——通用提示不会
            # 主动告诉用户 .xls 为什么不行、该怎么办。
            _die(
                ".xls 是旧版 Excel 格式，本工具读不了"
                "（.xls 是旧格式，请先在 Excel 里另存为 .xlsx 再试）。"
            )
        else:
            _die(
                f"只支持 .xlsx / .xlsm / .csv，收到的是 {suffix or '（无扩展名）'}"
            )

        print(f"[doc-convert] 已生成 {dst}")
        return 0
    except Exception as e:
        # 兜底：main() 必须有这一层。最现实的漏网是 wb.save() / CSV 写盘遇到
        # 只读目录、磁盘满时抛的裸 OSError，以及 openpyxl 读到损坏 zip 包时抛的
        # BadZipFile——冒泡出去都是一屏英文堆栈。同 doc_text.py / pdf_tables.py
        # 的纪律。SystemExit 继承 BaseException，主动报错退出不会被这里重新包一层。
        _die(f"处理过程中出错：{type(e).__name__}: {e}")
        return 2  # 不可达，只为类型完整


if __name__ == "__main__":
    raise SystemExit(main())
