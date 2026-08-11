#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""rows_to_xlsx.py — 把模型产出的 JSON/JSONL 装配成 Excel。

本 PR 全部质量门禁的落点。「PDF 表格转 Excel」和「票据批量转台账」共用它。

为什么装配一定要由脚本做、不能让模型直接写文件：
  1. 模型现写 Python 每次写得不一样，输出不稳定（总设计决策 #3 排除
     「零技能」方案的同一理由）。
  2. 更重要的是——确定性质检需要一个落点。「行列对不齐」「偷偷加了一列」
     「一半字段都没认出来」这些判断必须由代码执行，让模型自查等于没查。

「看不清」和「本来就没有」严格区分（这是本文件最容易被改坏的地方）：
  - 字段出现在 _存疑 里 → 写「⚠ 无法识别」+ 染黄底
  - 值是 null 但不在 _存疑 里 → 留空，什么都不写（票据上本来就没这项）
  混为一谈会制造大量假警报，用户三天就学会无视所有黄格子，标记随之失效。

数值靠 JSON 的类型区分：模型输出 128.5（number）就写成数值，输出 "128.5"
（string）就写成文本。「这是不是一个数」的判断交给看得懂上下文的模型，
写不写成数值的执行交给脚本——只有真数值才能被 =SUM() 算进去。
"""
import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

UNREADABLE_TEXT = "⚠ 无法识别"
YELLOW_HEX = "FFF2CC"
MAX_UNCERTAIN_RATIO_DEFAULT = 0.5
REVIEW_SHEET = "待核对"

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


def _die(msg: str) -> None:
    print(f"[doc-convert] 错误：{msg}", file=sys.stderr)
    raise SystemExit(2)


def load(path: Path, headers_arg: list[str] | None) -> tuple[list[str], list[dict], dict]:
    path = Path(path)
    if not path.is_file():
        _die(f"找不到输入文件「{path}」。请确认路径。")
    raw = path.read_text(encoding="utf-8")

    if path.suffix.lower() == ".jsonl":
        # JSONL 各行字段可能不齐（这正是它适合断点续跑的原因），靠首行推断
        # 表头会静默漏列——漏掉的那一列用户根本不会发现。所以强制显式指定。
        if not headers_arg:
            _die("读 .jsonl 必须用 --headers 显式给出表头。"
                 "各行字段可能不齐，靠第一行猜会静默漏列。")
        rows = []
        for i, line in enumerate(raw.splitlines(), start=1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError:
                _die(f"中间文件第 {i} 行不是合法 JSON，可能是上次跑到一半被打断了。"
                     "删掉这一行后重试即可，前面已识别的内容不会丢。")
            if not isinstance(obj, dict):
                # 评审实测：某一行合法 JSON 但不是对象（比如手滑写成了
                # [1,2,3]）会让下游 validate() 对着非 dict 做 `k in row`
                # 之类的操作，炸出英文 TypeError/AttributeError，兜底虽然
                # 接得住但诊断信息对用户毫无意义。这里提前用行号把问题
                # 钉死，比让它滚到后面变成一句看不懂的英文异常有用得多。
                _die(f"中间文件第 {i} 行不是一个对象（拿到的是 "
                     f"{type(obj).__name__}）。每一行必须是「字段名: 值」的"
                     "对象，请检查这一行的格式。")
            rows.append(obj)
        return list(headers_arg), rows, {}

    try:
        payload = json.loads(raw)
    except json.JSONDecodeError:
        _die(f"「{path.name}」不是合法的 JSON。")
    if not isinstance(payload, dict):
        # 评审实测：模型把 rows 数组直接当成整个文件（裸 `[{...}, {...}]`）
        # 是很自然的手滑——顶层缺了一层 {headers, rows} 包装。不提前拦住，
        # 下面 `payload.get("headers")` 会直接抛 AttributeError，被外层兜底
        # 接住后吐出一句用户看不懂的英文异常，等于白兜。
        _die(f"「{path.name}」顶层必须是一个包含 headers/rows 的对象，"
             f"但读到的是 {type(payload).__name__}。你是不是把 rows 数组"
             "直接当成了整个文件？请包一层：{\"headers\": [...], "
             "\"rows\": [...]}。")
    headers = headers_arg or payload.get("headers") or []
    return list(headers), payload.get("rows") or [], payload.get("meta") or {}


def _uncertain_fields(row: dict) -> set[str]:
    return {str(d.get("字段")) for d in (row.get("_存疑") or []) if d.get("字段")}


def validate(headers: list[str], rows: list[dict], max_ratio: float) -> None:
    if not headers:
        _die("没有表头，出不了表。请先确定这张表要哪几列。")
    if not rows:
        _die("一行数据都没有，不生成空表格。请确认输入内容。")

    known = set(headers)
    for i, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            # 评审实测：rows 里混进一个非对象元素（比如一个裸字符串）会让
            # 下面 `for k in row` 逐字符遍历那个字符串，报出的「表头里没有
            # 的字段」是一堆断成单字的假信息，完全误导用户去哪查。行号
            # 必须在这里就钉住，不能让它伪装成看似正常的「多余字段」错误。
            _die(f"第 {i} 行不是一个对象（拿到的是 {type(row).__name__}）。"
                 "每一行必须是「字段名: 值」的对象，请检查这一行的格式。")
        extra = [k for k in row if not k.startswith("_") and k not in known]
        if extra:
            _die(f"第 {i} 行出现了表头里没有的字段：{'、'.join(extra)}。"
                 "表头是跟用户确认过的，不能中途加列。请把它并进已有列，"
                 "或者先跟用户确认要不要加这一列。")
        # 「_存疑」里的字段名必须落在真实表头上——这是头号不变量（看不清
        # vs 本来没有）的另一半：如果字段名对不上（全角/半角括号、多个
        # 尾随空格、模型在存疑里写了简称……），build() 里 `name in unc` 的
        # 交集判断会静默找不到匹配，那一格既不染黄也不写「⚠ 无法识别」，
        # 还不计入下面的 50% 门禁——模型明明说了「这格没看清」，用户拿到
        # 的却是一个完全正常、没有任何标记的格子。同「偷加列」一个形状的
        # 校验，在这里当场拦住，不能指望 build() 兜底。
        bad = [f for f in _uncertain_fields(row) if f not in known]
        if bad:
            _die(f"第 {i} 行的 _存疑 里写了表头没有的字段：{'、'.join(bad)}。"
                 "存疑标记必须落在真实存在的列上，否则那一格不会被标黄，"
                 "用户看不出它有问题。请把字段名改成与表头完全一致。")

    uncertain = sum(len(_uncertain_fields(r) & known) for r in rows)
    total = len(rows) * len(headers)
    if total and uncertain / total > max_ratio:
        pct = round(uncertain / total * 100)
        _die(f"这批内容有 {pct}% 的字段没能认出来，超过一半，不生成表格。"
             "一份大半是问号的表，核对的工夫比重新录还多。"
             "建议把图片拍清楚些（光线充足、正对、别有折痕），或改用扫描件重试。")


INVALID_SHEET_CHARS = set("/\\[]:*?")  # Excel 工作表名禁用的字符
MAX_SHEET_NAME_LEN = 31  # Excel 硬限制，超过会在打开时静默截断或报错


def _validate_sheet_name(name: str) -> None:
    # 评审实测：`--sheet "2026/03 台账"` 这种带日期斜杠的名字，模型给得
    # 太自然了——不提前拦，openpyxl 会抛 `ValueError: Invalid character /
    # found in sheet title`，一句英文异常，用户不知道该改哪。
    bad = sorted(set(name) & INVALID_SHEET_CHARS)
    if bad:
        _die(f"工作表名「{name}」包含 Excel 不允许的字符：{' '.join(bad)}。"
             "Excel 表名不能含 / \\ [ ] : * ? 这几个符号，"
             "请换一个名字（比如把 / 换成 -）。")
    if len(name) > MAX_SHEET_NAME_LEN:
        _die(f"工作表名「{name}」长度 {len(name)} 超过 Excel 限制的 "
             f"{MAX_SHEET_NAME_LEN} 个字符，请换短一点的名字。")


def build(headers: list[str], rows: list[dict], meta: dict, sheet_name: str) -> Workbook:
    _validate_sheet_name(sheet_name)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(list(headers))
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"  # 滚动时表头不跑掉

    fill = PatternFill("solid", fgColor=YELLOW_HEX)
    review: list[tuple[str, str, str]] = []

    for row in rows:
        unc = _uncertain_fields(row)
        ws.append([None] * len(headers))
        r = ws.max_row
        for col, name in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=col)
            if name in unc:
                # 染黄底不是装饰：用户拿到 Excel 是拖着滚动条扫的，不会逐格看，
                # 颜色是唯一能在扫视中被捕捉到的信号。
                cell.value = UNREADABLE_TEXT
                cell.fill = fill
                continue
            v = row.get(name)
            if v is None:
                continue  # 本来就没有 → 留空，与「看不清」严格区分
            if isinstance(v, bool):
                cell.value = str(v)  # bool 是 int 的子类，不特判会写成 1
            elif isinstance(v, (int, float)):
                cell.value = v
            else:
                cell.value = str(v)
                if cell.data_type == "f":
                    # openpyxl 见到以 = 开头的字符串（比如模型抽出来的原文
                    # 恰好是 "=1+1" 这种备注）会自动把单元格当成公式。模型
                    # 抽出来的原文必须原样呈现，不能被 Excel 求值成结果或
                    # #NAME? 报错——这同样是「输出字符串就写成文本」这条
                    # 契约的一部分，只是被 openpyxl 的自动判型悄悄破坏了。
                    cell.data_type = "s"
        for d in row.get("_存疑") or []:
            review.append((str(row.get("_来源") or ""), str(d.get("字段") or ""),
                           str(d.get("原因") or "")))

    for col, name in enumerate(headers, start=1):
        # 收进列表再取 max：用 max(x, *generator) 的写法在只有表头行时
        # 会退化成 max(int) 抛 TypeError。这里虽然已经校验过 rows 非空，
        # 但别把正确性押在别处的校验上。
        lengths = [len(str(name)) * 2 + 2]
        lengths += [len(str(ws.cell(row=r, column=col).value or "")) + 2
                    for r in range(2, ws.max_row + 1)]
        ws.column_dimensions[get_column_letter(col)].width = min(max(lengths), 48)

    # 一个存疑都没有时不建这张表：一张空的「待核对」是噪音，
    # 会让用户以为有东西要核对。
    if review:
        rs = wb.create_sheet(REVIEW_SHEET)
        rs.append(["来源", "字段", "原因"])
        for c in rs[1]:
            c.font = Font(bold=True)
        for item in review:
            rs.append(list(item))
        for col, w in ((1, 32), (2, 16), (3, 60)):
            rs.column_dimensions[get_column_letter(col)].width = w

    if meta.get("标题"):
        wb.properties.title = str(meta["标题"])
    return wb


def main(argv: list[str] | None = None) -> int:
    try:
        ap = argparse.ArgumentParser(description="把模型产出的 JSON/JSONL 装配成 Excel")
        ap.add_argument("input", help="rows.json 或 rows.jsonl")
        ap.add_argument("-o", "--output", required=True, help="输出 .xlsx")
        ap.add_argument("--headers", nargs="+", help="表头（.jsonl 必填）")
        ap.add_argument("--sheet", default="数据", help="工作表名，默认「数据」")
        ap.add_argument("--max-uncertain-ratio", type=float,
                        default=MAX_UNCERTAIN_RATIO_DEFAULT,
                        help=f"存疑字段占比上限，默认 {MAX_UNCERTAIN_RATIO_DEFAULT}")
        args = ap.parse_args(argv)

        headers, rows, meta = load(Path(args.input), args.headers)
        # 先全部校验通过再写盘：任何一条不合格都整体不落盘，绝不留半成品文件
        # （同 PR 1 pdf_ops.py split 的两阶段做法）。
        validate(headers, rows, args.max_uncertain_ratio)

        wb = build(headers, rows, meta, args.sheet)
        dst = Path(args.output)
        try:
            dst.parent.mkdir(parents=True, exist_ok=True)
        except Exception:
            # 写盘调用单独兜底，不能让磁盘满/权限拒绝这类系统层异常
            # 裸露成 Traceback——同 pdf_render.py / doc_text.py / pdf_tables.py
            # 的纪律。
            _die(f"无法创建输出目录 {dst.parent}，请检查目录权限或磁盘空间。")
        try:
            wb.save(str(dst))
        except Exception:
            _die(f"写入 Excel 文件 {dst} 失败，请检查目标目录权限或磁盘空间。")

        print(f"已生成 {dst}（{len(rows)} 行 × {len(headers)} 列）")
        return 0
    except Exception as e:
        # 兜底：main() 必须有这一层，逐个函数自觉包 try 是不够的——任何未
        # 预期的异常（哪怕来自看起来无辜的一行代码，如 img.save() 遇到
        # 磁盘满/权限拒绝）都要转成中文错误，不能让裸 Traceback 打到用户
        # 面前。这是本 PR 的全局约束，同 pdf_render.py / doc_text.py /
        # pdf_tables.py 等脚本的纪律保持一致。
        _die(f"处理过程中出错：{type(e).__name__}: {e}")


if __name__ == "__main__":
    raise SystemExit(main())
